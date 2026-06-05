"""
CineMatch AI — Excel deliverables exporter.

Produces the human-readable Excel bundle that the brief requires
("את הדאטה עצמו אחרי המניפולציות שעשיתם … אפשר ZIP") on top of the
parquet/json artefacts used by the engine.

Output directory: data/excel_exports/
    catalog_full.xlsx          — every clean catalog row × 22 cols
    cleaning_log.xlsx          — per-source raw/clean/dropped counts
    evaluation_results.xlsx    — Spearman, Agreement@10, Precision@5, baselines
    trends.xlsx                — decade ratings, rising/declining genres, entropy
    awards_predictions.xlsx    — model output + tier per title
    data_dictionary.xlsx       — schema / engineered-feature explanations

Each sheet uses a brand-coloured header row (#2E4A7A), bold white text, frozen
first row, autofilter and auto-fit column widths.
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent
DATA = BASE / "data"
OUT = DATA / "excel_exports"
OUT.mkdir(parents=True, exist_ok=True)

BRAND_FILL = PatternFill(start_color="2E4A7A", end_color="2E4A7A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _style_sheet(ws, n_cols: int, n_rows: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = BRAND_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    if n_rows > 0 and n_cols > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
    # auto-fit (cap width to keep large columns sane)
    for c in range(1, n_cols + 1):
        col_letter = get_column_letter(c)
        max_len = len(str(ws.cell(row=1, column=c).value or ""))
        for r in range(2, min(n_rows + 2, 200)):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            ln = len(str(v))
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def _write_df(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        for name, df in sheets.items():
            df.to_excel(xl, sheet_name=name[:31], index=False)
        for name, df in sheets.items():
            ws = xl.book[name[:31]]
            _style_sheet(ws, n_cols=len(df.columns), n_rows=len(df))
    print(f"  ✅ {path.relative_to(BASE)}  ({sum(len(df) for df in sheets.values()):,} rows total)")


# ── 1. catalog_full ──────────────────────────────────────────────────────────

def export_catalog() -> None:
    cat = pd.read_parquet(DATA / "catalog.parquet")
    # bolt on award probability if available
    p = DATA / "award_predictions.parquet"
    if p.exists():
        preds = pd.read_parquet(p)
        cat = cat.merge(preds, on="title", how="left")
    _write_df(OUT / "catalog_full.xlsx", {"catalog": cat})


# ── 2. cleaning_log ──────────────────────────────────────────────────────────

def export_cleaning_log() -> None:
    log = json.load(open(DATA / "cleaning_log.json"))
    rows = [
        {"source": src, "raw_rows": v["raw"], "clean_rows": v["clean"],
         "dropped": v["dropped"], "drop_pct": v["drop_pct"]}
        for src, v in log["sources"].items()
    ]
    df_sources = pd.DataFrame(rows)
    df_totals = pd.DataFrame([{
        "total_raw": log["total_raw"],
        "total_clean": log["total_clean"],
        "pre_dedupe": log["pre_dedupe"],
        "post_dedupe": log["post_dedupe"],
        "deduped": log["deduped"],
        "drop_rate_overall_pct": log["drop_rate_overall"],
    }])
    _write_df(OUT / "cleaning_log.xlsx", {
        "by_source": df_sources,
        "totals": df_totals,
    })


# ── 3. evaluation_results ────────────────────────────────────────────────────

def export_evaluation() -> None:
    ev = json.load(open(DATA / "evaluation_results.json"))

    def _pivot(d: dict) -> pd.DataFrame:
        return pd.DataFrame([{"method": k, "value": v} for k, v in d.items()])

    sheets: dict[str, pd.DataFrame] = {}
    if "spearman" in ev:
        sheets["spearman"] = _pivot(ev["spearman"])
    if "agreement_at_10" in ev:
        sheets["agreement_at_10"] = _pivot(ev["agreement_at_10"])
    if "precision_at_5" in ev:
        sheets["precision_at_5"] = _pivot(ev["precision_at_5"])

    misc_rows = []
    if "anomaly_threshold_5pct" in ev:
        misc_rows.append({"key": "anomaly_threshold_5pct",
                          "value": ev["anomaly_threshold_5pct"]})
    if "best_score_stats" in ev and isinstance(ev["best_score_stats"], dict):
        for k, v in ev["best_score_stats"].items():
            misc_rows.append({"key": f"best_score.{k}", "value": v})
    if misc_rows:
        sheets["meta"] = pd.DataFrame(misc_rows)

    _write_df(OUT / "evaluation_results.xlsx", sheets)


# ── 4. trends ────────────────────────────────────────────────────────────────

def export_trends() -> None:
    p = DATA / "trends.json"
    if not p.exists():
        print("  ⚠ trends.json missing — skipping")
        return
    tr = json.load(open(p))
    sheets: dict[str, pd.DataFrame] = {}
    if "decade_stats" in tr:
        sheets["decade_stats"] = pd.DataFrame(tr["decade_stats"])
    if "top_rising_genres" in tr:
        sheets["rising_genres"] = pd.DataFrame(tr["top_rising_genres"])
    if "top_declining_genres" in tr:
        sheets["declining_genres"] = pd.DataFrame(tr["top_declining_genres"])
    if "genre_entropy_by_decade" in tr:
        sheets["entropy_by_decade"] = pd.DataFrame(tr["genre_entropy_by_decade"])
    if "headline" in tr and isinstance(tr["headline"], dict):
        sheets["headline"] = pd.DataFrame([tr["headline"]])
    _write_df(OUT / "trends.xlsx", sheets)


# ── 5. awards_predictions ────────────────────────────────────────────────────

def export_awards() -> None:
    p = DATA / "award_predictions.parquet"
    if not p.exists():
        print("  ⚠ award_predictions.parquet missing — skipping")
        return
    df = pd.read_parquet(p).sort_values("award_probability", ascending=False)

    def _tier(x: float) -> str:
        if x >= 0.75:
            return "🏆 Strong candidate"
        if x >= 0.50:
            return "⭐ Possible nominee"
        return "—"
    df["tier"] = df["award_probability"].map(_tier)

    sheets = {"all_predictions": df}
    ev_path = DATA / "award_model_eval.json"
    if ev_path.exists():
        rep = json.load(open(ev_path))
        rows = []
        for name, m in rep["models"].items():
            rows.append({
                "model": name,
                "cv_auc_mean": round(m["cv_auc_mean"], 4),
                "cv_auc_std":  round(m["cv_auc_std"], 4),
                "test_roc_auc": round(m["test_roc_auc"], 4),
                "test_avg_precision": round(m["test_avg_precision"], 4),
                "precision_at_10": round(m["precision_at_10"], 4),
                "recall_at_10": round(m["recall_at_10"], 4),
                "precision_at_50": round(m["precision_at_50"], 4),
            })
        sheets["model_eval"] = pd.DataFrame(rows)
    _write_df(OUT / "awards_predictions.xlsx", sheets)


# ── 6. data_dictionary ───────────────────────────────────────────────────────

DICTIONARY = [
    ("title", "string", "Show / movie title — primary key after dedupe"),
    ("language", "string", "ISO language code (en, he, …) from source feed"),
    ("start_year", "float", "First-broadcast / release year"),
    ("end_year", "float", "Final-broadcast year (NaN if ongoing)"),
    ("genres", "string", "Comma-separated genre list, title-cased"),
    ("rating", "float", "IMDb / TMDb 1–10 audience rating"),
    ("votes", "float", "Number of audience votes (popularity proxy)"),
    ("popularity", "float", "TMDb popularity index"),
    ("overview", "string", "Plot synopsis (TMDb / IMDb)"),
    ("poster_path", "string", "Relative TMDb poster URL fragment ('' = none)"),
    ("num_episodes", "float", "Total episode count if available"),
    ("num_seasons", "float", "Total season count if available"),
    ("source_dataset", "string", "Origin source feed (tmdb_tvs / disney_plus / imdb_all / imdb_top5000)"),
    ("decade", "float", "Start-year rounded down to nearest 10"),
    ("decade_str", "string", "Human-readable decade label (e.g. '2010s')"),
    ("genre_set_str", "string", "Pipe-joined genre set used by Jaccard"),
    ("rating_z", "float", "Z-scored rating (engineered)"),
    ("votes_z", "float", "Z-scored votes (engineered)"),
    ("start_year_z", "float", "Z-scored start year (engineered)"),
    ("popularity_z", "float", "Z-scored popularity (engineered)"),
    ("rating_bucket", "string", "Discrete rating tier label"),
    ("binge_fit_score", "float", "Display-only label — short-runtime + high-rating heuristic"),
    ("award_probability", "float", "Output of supervised LR/RF — P(Emmy or Oscar nominee)"),
]


def export_dictionary() -> None:
    df = pd.DataFrame(DICTIONARY, columns=["column", "type", "description"])
    notes = pd.DataFrame([
        {"key": "engineered_features",
         "note": "rating_z, votes_z, start_year_z, popularity_z are produced by data/load_catalog.py."},
        {"key": "jaccard_input",
         "note": "genre_set_str is split on '|' to form the discrete set used in engine/jaccard.py."},
        {"key": "embeddings",
         "note": "Plot synopsis embeddings are stored separately in data/embeddings.npy (11,013 × 384 float32, multilingual MiniLM)."},
        {"key": "award_model",
         "note": "Trained by engine/awards.py — see data/award_model_eval.json for CV/test metrics."},
    ])
    _write_df(OUT / "data_dictionary.xlsx", {"columns": df, "notes": notes})


def main() -> None:
    print(f"→ Exporting Excel deliverables to {OUT.relative_to(BASE)}/ …")
    export_catalog()
    export_cleaning_log()
    export_evaluation()
    export_trends()
    export_awards()
    export_dictionary()
    print("✅ All Excel exports done.")


if __name__ == "__main__":
    main()
